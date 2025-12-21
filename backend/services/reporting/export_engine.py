"""
Export Engine - Export snapshots to various formats
Supports CSV, XLSX, and PDF exports
"""

import csv
import io
import logging
from datetime import datetime
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def _xlsx_write_metadata(ws: "Worksheet", snapshot: dict[str, Any]) -> int:
    """Write metadata section to worksheet, return next row number."""
    from openpyxl.styles import Font

    ws["A1"] = "Snapshot Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Name:"
    ws["B2"] = snapshot.get("name", "Untitled")
    ws["A3"] = "Created:"
    ws["B3"] = datetime.fromtimestamp(snapshot["created_at"]).isoformat()
    ws["A4"] = "Created By:"
    ws["B4"] = snapshot.get("created_by", "Unknown")
    return 6


def _xlsx_write_summary(ws: "Worksheet", summary: dict[str, Any], start_row: int) -> int:
    """Write summary section to worksheet, return next row number."""
    from openpyxl.styles import Font

    ws[f"A{start_row}"] = "Summary"
    ws[f"A{start_row}"].font = Font(bold=True)
    row_num = start_row + 1
    for key, value in summary.items():
        ws[f"A{row_num}"] = key
        ws[f"B{row_num}"] = value
        row_num += 1
    return row_num + 1


def _xlsx_write_data(ws: "Worksheet", row_data: list[dict[str, Any]], start_row: int) -> int:
    """Write data rows to worksheet, return next row number."""
    from openpyxl.styles import Font, PatternFill

    if not row_data:
        return start_row

    headers = list(row_data[0].keys())
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    row_num = start_row + 1
    for row in row_data:
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num, value=row.get(header, ""))
        row_num += 1
    return row_num


def _xlsx_auto_column_widths(ws: "Worksheet") -> None:
    """Auto-adjust column widths based on content."""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > max_length:
                    max_length = cell_len
            except (AttributeError, TypeError, ValueError) as e:
                logger.debug(f"Could not calculate cell width: {e}")
                continue
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


class ExportEngine:
    """
    Export snapshots to different formats
    """

    def __init__(self):
        pass

    def export_to_csv(self, snapshot: dict[str, Any], include_summary: bool = True) -> bytes:
        """
        Export snapshot to CSV

        Args:
            snapshot: Snapshot document
            include_summary: Include summary at top

        Returns:
            CSV bytes
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Write metadata
        writer.writerow(["Snapshot Report"])
        writer.writerow(["Name:", snapshot.get("name", "Untitled")])
        writer.writerow(["Created:", datetime.fromtimestamp(snapshot["created_at"]).isoformat()])
        writer.writerow(["Created By:", snapshot.get("created_by", "Unknown")])
        writer.writerow([])

        # Write summary
        if include_summary and snapshot.get("summary"):
            writer.writerow(["Summary"])
            for key, value in snapshot["summary"].items():
                writer.writerow([key, value])
            writer.writerow([])

        # Write data
        row_data = snapshot.get("row_data", [])

        if row_data:
            # Get headers from first row
            headers = list(row_data[0].keys())
            writer.writerow(headers)

            # Write rows
            for row in row_data:
                writer.writerow([row.get(h, "") for h in headers])

        # Convert to bytes
        csv_bytes = output.getvalue().encode("utf-8")
        output.close()

        logger.info(f"✓ CSV export created: {len(row_data)} rows")
        return csv_bytes

    def export_to_xlsx(self, snapshot: dict[str, Any], include_summary: bool = True) -> bytes:
        """
        Export snapshot to Excel (XLSX)

        Requires: openpyxl
        """
        try:
            from openpyxl import Workbook
        except ImportError:
            raise ImportError("openpyxl is required for XLSX export")

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Write sections using helpers
        row_num = _xlsx_write_metadata(ws, snapshot)

        if include_summary and snapshot.get("summary"):
            row_num = _xlsx_write_summary(ws, snapshot["summary"], row_num)

        row_data = snapshot.get("row_data", [])
        _xlsx_write_data(ws, row_data, row_num)
        _xlsx_auto_column_widths(ws)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        xlsx_bytes = output.getvalue()
        output.close()

        logger.info(f"✓ XLSX export created: {len(row_data)} rows")
        return xlsx_bytes

    def export_to_json(self, snapshot: dict[str, Any]) -> bytes:
        """
        Export snapshot to JSON
        """
        import json

        # Remove MongoDB ObjectId if present
        snapshot_copy = {**snapshot}
        if "_id" in snapshot_copy:
            snapshot_copy["_id"] = str(snapshot_copy["_id"])

        json_bytes = json.dumps(snapshot_copy, indent=2, default=str).encode("utf-8")

        logger.info("✓ JSON export created")
        return json_bytes

    def get_export_filename(self, snapshot: dict[str, Any], format: str = "csv") -> str:
        """
        Generate export filename
        """
        name = snapshot.get("name", "report").replace(" ", "_")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{name}_{timestamp}.{format}"
