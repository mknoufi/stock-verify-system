"""
Advanced Report Service
Comprehensive report generation with real-time data, aggregations, and advanced filtering
"""

import csv
import io
import json
import logging
from datetime import datetime
from enum import Enum
from typing import Any, Optional

from pydantic import BaseModel, Field

from backend.utils.tracing import trace_report_generation, trace_span

logger = logging.getLogger(__name__)


class ReportFormat(str, Enum):
    JSON = "json"
    CSV = "csv"
    XLSX = "xlsx"
    PDF = "pdf"


class AggregationType(str, Enum):
    SUM = "sum"
    AVG = "avg"
    COUNT = "count"
    MIN = "min"
    MAX = "max"
    GROUP = "group"


class SortOrder(str, Enum):
    ASC = "asc"
    DESC = "desc"


class ColumnConfig(BaseModel):
    """Column configuration for dynamic table display."""

    field: str
    label: str
    visible: bool = True
    sortable: bool = True
    filterable: bool = True
    width: Optional[int] = None
    format: Optional[str] = None  # "date", "currency", "percentage", "number"
    aggregation: Optional[AggregationType] = None


class ReportFilters(BaseModel):
    """Advanced filter configuration for reports."""

    date_from: Optional[datetime] = None
    date_to: Optional[datetime] = None
    warehouse: Optional[str] = None
    floor: Optional[str] = None
    rack_id: Optional[str] = None
    category: Optional[str] = None
    user_id: Optional[str] = None
    session_id: Optional[str] = None
    verified: Optional[bool] = None
    status: Optional[str] = None
    variance_min: Optional[float] = None
    variance_max: Optional[float] = None
    item_code: Optional[str] = None
    search_query: Optional[str] = None


class ReportConfig(BaseModel):
    """Configuration for report generation."""

    report_type: str
    filters: Optional[ReportFilters] = None
    columns: Optional[list[ColumnConfig]] = None
    sort_by: Optional[str] = None
    sort_order: SortOrder = SortOrder.DESC
    page: int = Field(default=1, ge=1)
    page_size: int = Field(default=50, ge=1, le=1000)
    include_aggregations: bool = True
    include_summary: bool = True
    export_format: ReportFormat = ReportFormat.JSON


class ReportSummary(BaseModel):
    """Report summary with aggregated statistics."""

    total_records: int
    filtered_records: int
    aggregations: dict[str, Any]
    generated_at: datetime
    generation_time_ms: float
    filters_applied: dict[str, Any]
    report_type: str
    report_name: str


class AdvancedReportService:
    """Advanced report generation service with tracing support."""

    # Default column configurations for each report type
    REPORT_COLUMNS: dict[str, list[ColumnConfig]] = {
        "verified_items": [
            ColumnConfig(field="item_code", label="Item Code", sortable=True),
            ColumnConfig(field="item_name", label="Item Name", sortable=True),
            ColumnConfig(field="barcode", label="Barcode", sortable=True),
            ColumnConfig(field="category", label="Category", sortable=True, filterable=True),
            ColumnConfig(field="warehouse", label="Warehouse", sortable=True, filterable=True),
            ColumnConfig(field="floor", label="Floor", sortable=True, filterable=True),
            ColumnConfig(field="rack_id", label="Rack", sortable=True, filterable=True),
            ColumnConfig(field="stock_qty", label="ERP Qty", sortable=True, format="number"),
            ColumnConfig(field="counted_qty", label="Counted Qty", sortable=True, format="number"),
            ColumnConfig(field="variance", label="Variance", sortable=True, format="number"),
            ColumnConfig(
                field="variance_percentage",
                label="Variance %",
                sortable=True,
                format="percentage",
            ),
            ColumnConfig(field="mrp", label="MRP", sortable=True, format="currency"),
            ColumnConfig(field="verified", label="Verified", sortable=True),
            ColumnConfig(field="verified_by", label="Verified By", sortable=True),
            ColumnConfig(field="verified_at", label="Verified At", sortable=True, format="date"),
            ColumnConfig(field="counted_by", label="Counted By", sortable=True),
            ColumnConfig(field="counted_at", label="Counted At", sortable=True, format="date"),
            ColumnConfig(field="session_id", label="Session ID", visible=False),
            ColumnConfig(field="notes", label="Notes", visible=False),
        ],
        "session_summary": [
            ColumnConfig(field="session_id", label="Session ID"),
            ColumnConfig(field="rack_id", label="Rack"),
            ColumnConfig(field="floor", label="Floor"),
            ColumnConfig(field="username", label="User"),
            ColumnConfig(field="status", label="Status"),
            ColumnConfig(field="started_at", label="Started", format="date"),
            ColumnConfig(field="completed_at", label="Completed", format="date"),
            ColumnConfig(field="total_items", label="Total Items", format="number"),
            ColumnConfig(field="verified_items", label="Verified", format="number"),
            ColumnConfig(field="total_variance", label="Total Variance", format="number"),
            ColumnConfig(field="duration_minutes", label="Duration (min)", format="number"),
        ],
        "variance_analysis": [
            ColumnConfig(field="item_code", label="Item Code"),
            ColumnConfig(field="item_name", label="Item Name"),
            ColumnConfig(field="warehouse", label="Warehouse"),
            ColumnConfig(field="stock_qty", label="ERP Qty", format="number"),
            ColumnConfig(field="counted_qty", label="Counted", format="number"),
            ColumnConfig(field="variance", label="Variance", format="number"),
            ColumnConfig(field="variance_percentage", label="Variance %", format="percentage"),
            ColumnConfig(field="mrp", label="Unit Price", format="currency"),
            ColumnConfig(field="value_impact", label="Value Impact", format="currency"),
            ColumnConfig(field="status", label="Status"),
            ColumnConfig(field="risk_level", label="Risk Level"),
        ],
        "user_productivity": [
            ColumnConfig(field="username", label="User"),
            ColumnConfig(field="role", label="Role"),
            ColumnConfig(field="total_scans", label="Total Scans", format="number"),
            ColumnConfig(field="items_verified", label="Items Verified", format="number"),
            ColumnConfig(field="sessions_completed", label="Sessions", format="number"),
            ColumnConfig(field="avg_items_per_hour", label="Items/Hour", format="number"),
            ColumnConfig(field="accuracy_rate", label="Accuracy %", format="percentage"),
            ColumnConfig(field="last_activity", label="Last Activity", format="date"),
        ],
    }

    def __init__(self, db):
        self.db = db

    def get_column_config(self, report_type: str) -> list[ColumnConfig]:
        """Get default column configuration for a report type."""
        return self.REPORT_COLUMNS.get(report_type, [])

    def _build_base_query(self, filters: ReportFilters) -> dict[str, Any]:
        """Build base query from simple filter fields."""
        query: dict[str, Any] = {}

        if filters.verified is not None:
            query["verified"] = filters.verified

        if filters.warehouse:
            query["warehouse"] = filters.warehouse

        if filters.floor:
            query["floor"] = filters.floor

        if filters.rack_id:
            query["rack_id"] = filters.rack_id

        if filters.session_id:
            query["session_id"] = filters.session_id

        if filters.user_id:
            query["counted_by"] = filters.user_id

        if filters.item_code:
            query["item_code"] = {"$regex": filters.item_code, "$options": "i"}

        return query

    def _add_search_filter(self, query: dict[str, Any], search_query: Optional[str]) -> None:
        """Add search filter to query."""
        if search_query:
            query["$or"] = [
                {"item_code": {"$regex": search_query, "$options": "i"}},
                {"item_name": {"$regex": search_query, "$options": "i"}},
                {"barcode": {"$regex": search_query, "$options": "i"}},
            ]

    def _add_date_filter(
        self,
        query: dict[str, Any],
        date_from: Optional[datetime],
        date_to: Optional[datetime],
    ) -> None:
        """Add date range filter to query."""
        if date_from or date_to:
            date_filter: dict[str, Any] = {}
            if date_from:
                date_filter["$gte"] = date_from
            if date_to:
                date_filter["$lte"] = date_to
            query["counted_at"] = date_filter

    def _add_variance_filter(
        self,
        query: dict[str, Any],
        variance_min: Optional[float],
        variance_max: Optional[float],
    ) -> None:
        """Add variance range filter to query."""
        if variance_min is not None or variance_max is not None:
            variance_filter: dict[str, Any] = {}
            if variance_min is not None:
                variance_filter["$gte"] = variance_min
            if variance_max is not None:
                variance_filter["$lte"] = variance_max
            query["variance"] = variance_filter

    def _get_items_projection(self) -> dict[str, Any]:
        """Get projection for verified items query."""
        return {
            "_id": 0,
            "id": 1,
            "item_code": 1,
            "item_name": 1,
            "barcode": 1,
            "category": 1,
            "warehouse": 1,
            "floor": 1,
            "rack_id": 1,
            "stock_qty": 1,
            "counted_qty": 1,
            "variance": 1,
            "variance_percentage": {
                "$cond": {
                    "if": {"$eq": ["$stock_qty", 0]},
                    "then": 0,
                    "else": {
                        "$multiply": [
                            {"$divide": ["$variance", {"$abs": "$stock_qty"}]},
                            100,
                        ]
                    },
                }
            },
            "mrp": {"$ifNull": ["$mrp_erp", 0]},
            "verified": 1,
            "verified_by": 1,
            "verified_at": 1,
            "counted_by": 1,
            "counted_at": 1,
            "session_id": 1,
            "notes": 1,
            "status": 1,
            "approval_status": 1,
        }

    @trace_report_generation("verified_items")
    async def generate_verified_items_report(self, config: ReportConfig) -> dict[str, Any]:
        """Generate report of verified items with full details."""
        start_time = datetime.utcnow()
        filters = config.filters or ReportFilters()

        # Build query using helpers
        query = self._build_base_query(filters)
        self._add_search_filter(query, filters.search_query)
        self._add_date_filter(query, filters.date_from, filters.date_to)
        self._add_variance_filter(query, filters.variance_min, filters.variance_max)

        # Get counts
        total_records = await self.db.count_lines.count_documents({})
        filtered_records = await self.db.count_lines.count_documents(query)

        # Build sort
        sort_field = config.sort_by or "counted_at"
        sort_direction = -1 if config.sort_order == SortOrder.DESC else 1

        # Pagination
        skip = (config.page - 1) * config.page_size

        # Execute query
        pipeline = [
            {"$match": query},
            {"$sort": {sort_field: sort_direction}},
            {"$skip": skip},
            {"$limit": config.page_size},
            {"$project": self._get_items_projection()},
        ]

        with trace_span("mongodb.count_lines.aggregate", {"query": str(query)[:200]}):
            data = await self.db.count_lines.aggregate(pipeline).to_list(config.page_size)

        # Calculate aggregations if requested
        aggregations = {}
        if config.include_aggregations:
            with trace_span("calculate_aggregations"):
                agg_pipeline = [
                    {"$match": query},
                    {
                        "$group": {
                            "_id": None,
                            "total_items": {"$sum": 1},
                            "total_variance": {"$sum": "$variance"},
                            "avg_variance": {"$avg": "$variance"},
                            "total_value": {
                                "$sum": {
                                    "$multiply": [
                                        "$counted_qty",
                                        {"$ifNull": ["$mrp_erp", 0]},
                                    ]
                                }
                            },
                            "variance_value": {
                                "$sum": {
                                    "$multiply": [
                                        "$variance",
                                        {"$ifNull": ["$mrp_erp", 0]},
                                    ]
                                }
                            },
                            "verified_count": {
                                "$sum": {"$cond": [{"$eq": ["$verified", True]}, 1, 0]}
                            },
                            "positive_variance": {
                                "$sum": {"$cond": [{"$gt": ["$variance", 0]}, "$variance", 0]}
                            },
                            "negative_variance": {
                                "$sum": {"$cond": [{"$lt": ["$variance", 0]}, "$variance", 0]}
                            },
                        }
                    },
                ]
                agg_result = await self.db.count_lines.aggregate(agg_pipeline).to_list(1)
                if agg_result:
                    aggregations = agg_result[0]
                    aggregations.pop("_id", None)

        end_time = datetime.utcnow()
        generation_time_ms = (end_time - start_time).total_seconds() * 1000

        # Build response
        columns = config.columns or self.get_column_config("verified_items")

        return {
            "success": True,
            "data": data,
            "columns": [col.model_dump() for col in columns],
            "summary": ReportSummary(
                total_records=total_records,
                filtered_records=filtered_records,
                aggregations=aggregations,
                generated_at=end_time,
                generation_time_ms=generation_time_ms,
                filters_applied=filters.model_dump(exclude_none=True),
                report_type="verified_items",
                report_name="Verified Items Report",
            ).model_dump(),
            "pagination": {
                "page": config.page,
                "page_size": config.page_size,
                "total_pages": (filtered_records + config.page_size - 1) // config.page_size,
                "has_next": skip + config.page_size < filtered_records,
                "has_prev": config.page > 1,
            },
        }

    @trace_report_generation("session_summary")
    async def generate_session_summary_report(self, config: ReportConfig) -> dict[str, Any]:
        """Generate session summary report with aggregated data."""
        start_time = datetime.utcnow()
        filters = config.filters or ReportFilters()

        query: dict[str, Any] = {}

        if filters.status:
            query["status"] = filters.status

        if filters.user_id:
            query["user_id"] = filters.user_id

        if filters.floor:
            query["floor"] = filters.floor

        if filters.date_from or filters.date_to:
            date_filter: dict[str, Any] = {}
            if filters.date_from:
                date_filter["$gte"] = filters.date_from
            if filters.date_to:
                date_filter["$lte"] = filters.date_to
            query["started_at"] = date_filter

        total_records = await self.db.sessions.count_documents({})
        filtered_records = await self.db.sessions.count_documents(query)

        sort_field = config.sort_by or "started_at"
        sort_direction = -1 if config.sort_order == SortOrder.DESC else 1
        skip = (config.page - 1) * config.page_size

        pipeline = [
            {"$match": query},
            {
                "$lookup": {
                    "from": "count_lines",
                    "localField": "id",
                    "foreignField": "session_id",
                    "as": "lines",
                }
            },
            {
                "$lookup": {
                    "from": "users",
                    "localField": "user_id",
                    "foreignField": "username",
                    "as": "user_info",
                }
            },
            {
                "$project": {
                    "_id": 0,
                    "session_id": "$id",
                    "rack_id": 1,
                    "floor": 1,
                    "username": {
                        "$ifNull": [
                            {"$arrayElemAt": ["$user_info.username", 0]},
                            "$user_id",
                        ]
                    },
                    "status": 1,
                    "started_at": 1,
                    "completed_at": 1,
                    "total_items": {"$size": "$lines"},
                    "verified_items": {
                        "$size": {
                            "$filter": {
                                "input": "$lines",
                                "as": "line",
                                "cond": {"$eq": ["$$line.verified", True]},
                            }
                        }
                    },
                    "total_variance": {"$sum": "$lines.variance"},
                    "duration_minutes": {
                        "$cond": {
                            "if": {"$and": ["$started_at", "$completed_at"]},
                            "then": {
                                "$divide": [
                                    {"$subtract": ["$completed_at", "$started_at"]},
                                    60000,
                                ]
                            },
                            "else": None,
                        }
                    },
                }
            },
            {"$sort": {sort_field: sort_direction}},
            {"$skip": skip},
            {"$limit": config.page_size},
        ]

        data = await self.db.sessions.aggregate(pipeline).to_list(config.page_size)

        end_time = datetime.utcnow()
        generation_time_ms = (end_time - start_time).total_seconds() * 1000

        columns = config.columns or self.get_column_config("session_summary")

        return {
            "success": True,
            "data": data,
            "columns": [col.model_dump() for col in columns],
            "summary": ReportSummary(
                total_records=total_records,
                filtered_records=filtered_records,
                aggregations={},
                generated_at=end_time,
                generation_time_ms=generation_time_ms,
                filters_applied=filters.model_dump(exclude_none=True),
                report_type="session_summary",
                report_name="Session Summary Report",
            ).model_dump(),
            "pagination": {
                "page": config.page,
                "page_size": config.page_size,
                "total_pages": (filtered_records + config.page_size - 1) // config.page_size,
                "has_next": skip + config.page_size < filtered_records,
                "has_prev": config.page > 1,
            },
        }

    @trace_report_generation("variance_analysis")
    async def generate_variance_analysis_report(self, config: ReportConfig) -> dict[str, Any]:
        """Generate variance analysis report with risk levels."""
        start_time = datetime.utcnow()
        filters = config.filters or ReportFilters()

        # Only include items with non-zero variance
        query: dict[str, Any] = {"variance": {"$ne": 0}}

        if filters.warehouse:
            query["warehouse"] = filters.warehouse

        if filters.category:
            query["category"] = filters.category

        if filters.date_from or filters.date_to:
            date_filter: dict[str, Any] = {}
            if filters.date_from:
                date_filter["$gte"] = filters.date_from
            if filters.date_to:
                date_filter["$lte"] = filters.date_to
            query["counted_at"] = date_filter

        total_records = await self.db.count_lines.count_documents({"variance": {"$ne": 0}})
        filtered_records = await self.db.count_lines.count_documents(query)

        sort_field = config.sort_by or "variance"
        sort_direction = -1 if config.sort_order == SortOrder.DESC else 1
        skip = (config.page - 1) * config.page_size

        pipeline = [
            {"$match": query},
            {
                "$project": {
                    "_id": 0,
                    "item_code": 1,
                    "item_name": 1,
                    "warehouse": 1,
                    "stock_qty": 1,
                    "counted_qty": 1,
                    "variance": 1,
                    "variance_percentage": {
                        "$cond": {
                            "if": {"$eq": ["$stock_qty", 0]},
                            "then": 100,
                            "else": {
                                "$multiply": [
                                    {
                                        "$divide": [
                                            {"$abs": "$variance"},
                                            {"$abs": "$stock_qty"},
                                        ]
                                    },
                                    100,
                                ]
                            },
                        }
                    },
                    "mrp": {"$ifNull": ["$mrp_erp", 0]},
                    "value_impact": {"$multiply": ["$variance", {"$ifNull": ["$mrp_erp", 0]}]},
                    "status": 1,
                    "risk_level": {
                        "$switch": {
                            "branches": [
                                {
                                    "case": {
                                        "$or": [
                                            {"$gte": [{"$abs": "$variance"}, 100]},
                                            {
                                                "$gte": [
                                                    {
                                                        "$abs": {
                                                            "$multiply": [
                                                                "$variance",
                                                                {
                                                                    "$ifNull": [
                                                                        "$mrp_erp",
                                                                        0,
                                                                    ]
                                                                },
                                                            ]
                                                        }
                                                    },
                                                    10000,
                                                ]
                                            },
                                        ]
                                    },
                                    "then": "Critical",
                                },
                                {
                                    "case": {
                                        "$or": [
                                            {"$gte": [{"$abs": "$variance"}, 50]},
                                            {
                                                "$gte": [
                                                    {
                                                        "$abs": {
                                                            "$multiply": [
                                                                "$variance",
                                                                {
                                                                    "$ifNull": [
                                                                        "$mrp_erp",
                                                                        0,
                                                                    ]
                                                                },
                                                            ]
                                                        }
                                                    },
                                                    5000,
                                                ]
                                            },
                                        ]
                                    },
                                    "then": "High",
                                },
                                {
                                    "case": {"$gte": [{"$abs": "$variance"}, 10]},
                                    "then": "Medium",
                                },
                            ],
                            "default": "Low",
                        }
                    },
                }
            },
            {"$sort": {sort_field: sort_direction}},
            {"$skip": skip},
            {"$limit": config.page_size},
        ]

        data = await self.db.count_lines.aggregate(pipeline).to_list(config.page_size)

        # Calculate aggregations
        aggregations = {}
        if config.include_aggregations:
            agg_pipeline = [
                {"$match": query},
                {
                    "$group": {
                        "_id": None,
                        "total_variance_items": {"$sum": 1},
                        "total_variance_qty": {"$sum": "$variance"},
                        "total_value_impact": {
                            "$sum": {"$multiply": ["$variance", {"$ifNull": ["$mrp_erp", 0]}]}
                        },
                        "positive_variance_count": {
                            "$sum": {"$cond": [{"$gt": ["$variance", 0]}, 1, 0]}
                        },
                        "negative_variance_count": {
                            "$sum": {"$cond": [{"$lt": ["$variance", 0]}, 1, 0]}
                        },
                    }
                },
            ]
            agg_result = await self.db.count_lines.aggregate(agg_pipeline).to_list(1)
            if agg_result:
                aggregations = agg_result[0]
                aggregations.pop("_id", None)

        end_time = datetime.utcnow()
        generation_time_ms = (end_time - start_time).total_seconds() * 1000

        columns = config.columns or self.get_column_config("variance_analysis")

        return {
            "success": True,
            "data": data,
            "columns": [col.model_dump() for col in columns],
            "summary": ReportSummary(
                total_records=total_records,
                filtered_records=filtered_records,
                aggregations=aggregations,
                generated_at=end_time,
                generation_time_ms=generation_time_ms,
                filters_applied=filters.model_dump(exclude_none=True),
                report_type="variance_analysis",
                report_name="Variance Analysis Report",
            ).model_dump(),
            "pagination": {
                "page": config.page,
                "page_size": config.page_size,
                "total_pages": (filtered_records + config.page_size - 1) // config.page_size,
                "has_next": skip + config.page_size < filtered_records,
                "has_prev": config.page > 1,
            },
        }

    async def export_to_csv(self, data: list[dict], columns: list[ColumnConfig]) -> str:
        """Export data to CSV format."""
        output = io.StringIO()
        visible_columns = [col for col in columns if col.visible]
        fieldnames = [col.field for col in visible_columns]
        headers = {col.field: col.label for col in visible_columns}

        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
        writer.writerow(headers)

        for row in data:
            sanitized_row = {}
            for field in fieldnames:
                value = row.get(field, "")
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                elif isinstance(value, datetime):
                    value = value.isoformat()
                sanitized_row[field] = value
            writer.writerow(sanitized_row)

        return output.getvalue()

    async def export_to_xlsx(self, data: list[dict], columns: list[ColumnConfig]) -> bytes:
        """Export data to XLSX format."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl required for Excel export")

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Report")

        visible_columns = [col for col in columns if col.visible]

        # Write headers
        for col_idx, col in enumerate(visible_columns, 1):
            ws.cell(row=1, column=col_idx, value=col.label)

        # Write data
        for row_idx, row in enumerate(data, 2):
            for col_idx, col in enumerate(visible_columns, 1):
                value = row.get(col.field, "")
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                elif isinstance(value, datetime):
                    value = value.isoformat()
                ws.cell(row=row_idx, column=col_idx, value=value)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


# Report type dispatcher
REPORT_GENERATORS = {
    "verified_items": "generate_verified_items_report",
    "session_summary": "generate_session_summary_report",
    "variance_analysis": "generate_variance_analysis_report",
}
