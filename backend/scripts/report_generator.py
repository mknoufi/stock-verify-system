"""
Report Generation Module
Generates comprehensive reports in multiple formats (Excel, CSV, PDF)
"""

import io
import logging

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generate various reports for stock verification"""

    @staticmethod
    def generate_session_summary_excel(session_data: dict, count_lines: list[dict]) -> bytes:
        """Generate comprehensive session summary in Excel format"""

        output = io.BytesIO()
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # 1. Summary Sheet
        ws_summary = wb.create_sheet("Session Summary")
        ReportGenerator._create_summary_sheet(ws_summary, session_data, count_lines)

        # 2. Detailed Counts Sheet
        ws_details = wb.create_sheet("Count Details")
        ReportGenerator._create_details_sheet(ws_details, count_lines)

        # 3. Variance Analysis Sheet
        ws_variance = wb.create_sheet("Variance Analysis")
        ReportGenerator._create_variance_sheet(ws_variance, count_lines)

        # 4. Aging Stock Sheet
        ws_aging = wb.create_sheet("Aging Stock")
        ReportGenerator._create_aging_stock_sheet(ws_aging, count_lines)

        # 5. Unmatched Items Sheet
        ws_unmatched = wb.create_sheet("Unmatched Items")
        ReportGenerator._create_unmatched_sheet(ws_unmatched, session_data, count_lines)

        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def _create_summary_sheet(ws, session_data: dict, count_lines: list[dict]):
        """Create session summary sheet"""
        # Header
        ws["A1"] = "LAVANYA E-MART - STOCK VERIFICATION SUMMARY"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        ws.merge_cells("A1:D1")

        # Session Info
        row = 3
        info_data = [
            ["Session ID:", session_data.get("id", "")],
            ["Warehouse:", session_data.get("warehouse", "")],
            ["Staff:", session_data.get("staff_name", "")],
            ["Status:", session_data.get("status", "")],
            ["Started:", str(session_data.get("started_at", ""))],
            ["Total Items:", str(session_data.get("total_items", 0))],
            ["Total Variance:", str(session_data.get("total_variance", 0))],
        ]

        for label, value in info_data:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

        # Statistics
        row += 2
        ws[f"A{row}"] = "STATISTICS"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        row += 1

        # Calculate stats
        total_items = len(count_lines)
        with_variance = len(
            [line_entry for line_entry in count_lines if line_entry.get("variance", 0) != 0]
        )
        positive_variance = sum(
            line_entry.get("variance", 0)
            for line_entry in count_lines
            if line_entry.get("variance", 0) > 0
        )
        negative_variance = sum(
            line_entry.get("variance", 0)
            for line_entry in count_lines
            if line_entry.get("variance", 0) < 0
        )

        stats_data = [
            ["Total Items Counted:", total_items],
            ["Items with Variance:", with_variance],
            ["Items Without Variance:", total_items - with_variance],
            ["Positive Variance (Excess):", f"{positive_variance:.2f}"],
            ["Negative Variance (Shortage):", f"{negative_variance:.2f}"],
            ["Net Variance:", f"{positive_variance + negative_variance:.2f}"],
        ]

        for label, value in stats_data:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

    @staticmethod
    def _create_details_sheet(ws, count_lines: list[dict]):
        """Create detailed count lines sheet"""
        # Headers
        headers = [
            "Item Code",
            "Item Name",
            "Barcode",
            "ERP Qty",
            "Counted Qty",
            "Variance",
            "Variance %",
            "Reason",
            "Remark",
            "Counted By",
            "Status",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Data
        for row, line in enumerate(count_lines, 2):
            erp_qty = line.get("erp_qty", 0)
            variance_pct = (line.get("variance", 0) / erp_qty * 100) if erp_qty > 0 else 0

            data = [
                line.get("item_code", ""),
                line.get("item_name", ""),
                line.get("barcode", ""),
                erp_qty,
                line.get("counted_qty", 0),
                line.get("variance", 0),
                f"{variance_pct:.2f}%",
                line.get("variance_reason", ""),
                line.get("remark", ""),
                line.get("counted_by", ""),
                line.get("status", "pending"),
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row, col, value)

                # Highlight variance
                if col == 6:  # Variance column
                    if value > 0:
                        cell.fill = PatternFill(
                            start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"
                        )
                    elif value < 0:
                        cell.fill = PatternFill(
                            start_color="E0E0FF", end_color="E0E0FF", fill_type="solid"
                        )

        # Auto-adjust columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = 15

    @staticmethod
    def _create_variance_sheet(ws, count_lines: list[dict]):
        """Create variance analysis sheet"""
        # Only items with variance
        variance_lines = [
            line_entry for line_entry in count_lines if line_entry.get("variance", 0) != 0
        ]

        # Headers
        headers = [
            "Item Name",
            "Barcode",
            "ERP Qty",
            "Counted Qty",
            "Variance",
            "Variance %",
            "Value Impact",
            "Reason",
            "Priority",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF5252", end_color="FF5252", fill_type="solid")

        # Data
        for row, line in enumerate(
            sorted(variance_lines, key=lambda x: abs(x.get("variance", 0)), reverse=True),
            2,
        ):
            erp_qty = line.get("erp_qty", 0)
            variance = line.get("variance", 0)
            variance_pct = (variance / erp_qty * 100) if erp_qty > 0 else 0
            mrp = line.get("mrp_erp", 0)
            value_impact = variance * mrp

            # Determine priority
            if abs(variance_pct) > 20:
                priority = "HIGH"
            elif abs(variance_pct) > 10:
                priority = "MEDIUM"
            else:
                priority = "LOW"

            data = [
                line.get("item_name", ""),
                line.get("barcode", ""),
                erp_qty,
                line.get("counted_qty", 0),
                variance,
                f"{variance_pct:.2f}%",
                f"₹{value_impact:.2f}",
                line.get("variance_reason", ""),
                priority,
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row, col, value)

                # Color code by priority
                if col == 9:
                    if value == "HIGH":
                        cell.fill = PatternFill(
                            start_color="FF0000", end_color="FF0000", fill_type="solid"
                        )
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif value == "MEDIUM":
                        cell.fill = PatternFill(
                            start_color="FFA500", end_color="FFA500", fill_type="solid"
                        )

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = 15

    @staticmethod
    def _create_aging_stock_sheet(ws, count_lines: list[dict]):
        """Create aging stock analysis sheet"""
        from barcode_analyzer import BarcodeAnalyzer

        # Analyze all barcodes
        aging_items = []
        for line in count_lines:
            barcode = line.get("barcode", "")
            analysis = BarcodeAnalyzer.analyze_barcode(barcode)

            if analysis["is_aging_stock"]:
                aging_items.append({**line, "analysis": analysis})

        # Headers
        headers = [
            "Item Name",
            "Barcode",
            "Stock Qty",
            "Category",
            "Age (Months)",
            "Recommended Discount",
            "Action Required",
            "Priority",
            "Recommendations",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")

        # Data
        for row, item in enumerate(
            sorted(aging_items, key=lambda x: x["analysis"]["age_months"], reverse=True),
            2,
        ):
            analysis = item["analysis"]

            data = [
                item.get("item_name", ""),
                item.get("barcode", ""),
                item.get("counted_qty", 0),
                analysis["category"],
                analysis["age_months"],
                f"{analysis['recommended_discount']}%",
                analysis["action_required"],
                analysis["priority"],
                "; ".join(analysis["recommendations"][:3]),  # Top 3 recommendations
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row, col, value)

                # Color code by priority
                if analysis["priority"] == "HIGH":
                    cell.fill = PatternFill(
                        start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                    )

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = 18

    @staticmethod
    def _create_unmatched_sheet(ws, session_data: dict, count_lines: list[dict]):
        """Create unmatched items sheet (in ERP but not counted)"""
        # This will be populated when we have full ERP item list
        # For now, show template

        headers = [
            "Item Code",
            "Item Name",
            "Barcode",
            "ERP Qty",
            "MRP",
            "Category",
            "Status",
            "Action Required",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")

        ws["A2"] = "NOTE: Unmatched items require full ERP vs Counted comparison"
        ws["A2"].font = Font(italic=True, color="666666")
        ws.merge_cells("A2:H2")

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = 15

    @staticmethod
    def generate_csv(data: list[dict], columns: list[str]) -> str:
        """Generate CSV from data"""
        df = pd.DataFrame(data)
        if columns:
            df = df[columns]
        return df.to_csv(index=False)

    @staticmethod
    def generate_variance_summary_csv(count_lines: list[dict]) -> str:
        """Generate variance summary CSV"""
        variance_data = []
        for line in count_lines:
            if line.get("variance", 0) != 0:
                variance_data.append(
                    {
                        "Item_Code": line.get("item_code"),
                        "Item_Name": line.get("item_name"),
                        "Barcode": line.get("barcode"),
                        "ERP_Qty": line.get("erp_qty"),
                        "Counted_Qty": line.get("counted_qty"),
                        "Variance": line.get("variance"),
                        "Reason": line.get("variance_reason"),
                        "Counted_By": line.get("counted_by"),
                    }
                )

        df = pd.DataFrame(variance_data)
        return df.to_csv(index=False)
