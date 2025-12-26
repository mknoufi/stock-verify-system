"""
Report Generation API - Multiple report types with filtering and export
Supports Stock Summary, Variance, User Activity, Session History, and Audit Trail reports
"""

import csv
import io
import json
import logging
from datetime import date, datetime
from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from backend.auth.dependencies import get_current_user, require_role
from backend.db.runtime import get_db

logger = logging.getLogger(__name__)

report_generation_router = APIRouter(prefix="/reports", tags=["Reports"])


# Report Types
REPORT_TYPES = {
    "stock_summary": {
        "name": "Stock Summary Report",
        "description": "Overview of stock levels, values, and verification status",
        "collection": "erp_items",
    },
    "variance_report": {
        "name": "Variance Report",
        "description": "Items with discrepancies between expected and counted quantities",
        "collection": "verification_records",
    },
    "user_activity": {
        "name": "User Activity Report",
        "description": "User actions, sessions, and productivity metrics",
        "collection": "audit_logs",
    },
    "session_history": {
        "name": "Session History Report",
        "description": "Verification session details and outcomes",
        "collection": "verification_sessions",
    },
    "audit_trail": {
        "name": "Audit Trail Report",
        "description": "Complete audit log of all system actions",
        "collection": "audit_logs",
    },
}


# Request/Response Models
class ReportFilter(BaseModel):
    date_from: Optional[date] = None
    date_to: Optional[date] = None
    warehouse: Optional[str] = None
    user_id: Optional[str] = None
    status: Optional[str] = None
    floor: Optional[str] = None
    category: Optional[str] = None


class ReportRequest(BaseModel):
    report_type: str
    filters: Optional[ReportFilter] = None
    format: str = Field(default="json", pattern="^(json|csv|xlsx)$")
    include_summary: bool = True


class ReportSummary(BaseModel):
    total_records: int
    generated_at: str
    filters_applied: dict[str, Any]
    report_type: str
    report_name: str


class ReportResponse(BaseModel):
    summary: ReportSummary
    data: list[dict[str, Any]]


# Helper Functions
def build_date_filter(
    date_from: Optional[date], date_to: Optional[date]
) -> Optional[dict[str, Any]]:
    """Build MongoDB date range filter."""
    date_filter: dict[str, Any] = {}

    if date_from:
        date_filter["$gte"] = datetime.combine(date_from, datetime.min.time())

    if date_to:
        date_filter["$lte"] = datetime.combine(date_to, datetime.max.time())

    return date_filter if date_filter else None


def sanitize_for_csv(value: Any) -> str:
    """Sanitize value for CSV export."""
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value)
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _format_xlsx_cell_value(value: Any) -> Any:
    """Format a value for Excel cell."""
    if isinstance(value, (dict, list)):
        return json.dumps(value)
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def _write_xlsx_headers(ws: Any, headers: list[str]) -> None:
    """Write header row to Excel worksheet."""
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)


def _write_xlsx_data(ws: Any, data: list[dict], headers: list[str]) -> None:
    """Write data rows to Excel worksheet."""
    for row_idx, row in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row.get(header)
            ws.cell(row=row_idx, column=col_idx, value=_format_xlsx_cell_value(value))


async def generate_stock_summary(db, filters: ReportFilter) -> list[dict]:
    """Generate stock summary report data."""
    query: dict[str, Any] = {}

    if filters.warehouse:
        query["warehouse"] = filters.warehouse
    if filters.floor:
        query["floor"] = filters.floor
    if filters.category:
        query["category"] = filters.category

    pipeline = [
        {"$match": query},
        {
            "$lookup": {
                "from": "verification_records",
                "localField": "item_code",
                "foreignField": "item_code",
                "as": "verifications",
            }
        },
        {
            "$project": {
                "_id": 0,
                "item_code": 1,
                "item_name": 1,
                "category": 1,
                "warehouse": 1,
                "floor": 1,
                "stock_qty": 1,
                "price": {"$ifNull": ["$price", 0]},
                "stock_value": {
                    "$multiply": ["$stock_qty", {"$ifNull": ["$price", 0]}]
                },
                "verification_count": {"$size": "$verifications"},
                "last_verified": {"$max": "$verifications.created_at"},
                "is_verified": {"$gt": [{"$size": "$verifications"}, 0]},
            }
        },
        {"$sort": {"item_code": 1}},
        {"$limit": 10000},
    ]

    return await db.erp_items.aggregate(pipeline).to_list(10000)


async def generate_variance_report(db, filters: ReportFilter) -> list[dict]:
    """Generate variance report data."""
    query: dict[str, Any] = {"variance": {"$ne": 0}}

    if filters.warehouse:
        query["warehouse"] = filters.warehouse
    if filters.status:
        query["status"] = filters.status
    if filters.user_id:
        query["scanned_by"] = filters.user_id

    date_filter = build_date_filter(filters.date_from, filters.date_to)
    if date_filter:
        query["created_at"] = date_filter

    pipeline = [
        {"$match": query},
        {
            "$lookup": {
                "from": "erp_items",
                "localField": "item_code",
                "foreignField": "item_code",
                "as": "item_info",
            }
        },
        {"$unwind": {"path": "$item_info", "preserveNullAndEmptyArrays": True}},
        {
            "$project": {
                "_id": 0,
                "item_code": 1,
                "item_name": {"$ifNull": ["$item_info.item_name", "Unknown"]},
                "expected_qty": 1,
                "counted_qty": 1,
                "variance": 1,
                "variance_percentage": {
                    "$cond": {
                        "if": {"$eq": ["$expected_qty", 0]},
                        "then": 100,
                        "else": {
                            "$multiply": [
                                {"$divide": ["$variance", {"$abs": "$expected_qty"}]},
                                100,
                            ]
                        },
                    }
                },
                "status": 1,
                "scanned_by": 1,
                "warehouse": 1,
                "location": 1,
                "created_at": 1,
                "approved_by": 1,
                "approved_at": 1,
            }
        },
        {"$sort": {"variance_percentage": -1}},
        {"$limit": 10000},
    ]

    return await db.verification_records.aggregate(pipeline).to_list(10000)


async def generate_user_activity_report(db, filters: ReportFilter) -> list[dict]:
    """Generate user activity report data."""
    query: dict[str, Any] = {}

    if filters.user_id:
        query["user_id"] = filters.user_id

    date_filter = build_date_filter(filters.date_from, filters.date_to)
    if date_filter:
        query["timestamp"] = date_filter

    pipeline = [
        {"$match": query},
        {
            "$group": {
                "_id": "$user_id",
                "total_actions": {"$sum": 1},
                "scans": {"$sum": {"$cond": [{"$eq": ["$action", "scan"]}, 1, 0]}},
                "verifications": {
                    "$sum": {"$cond": [{"$eq": ["$action", "verify"]}, 1, 0]}
                },
                "approvals": {
                    "$sum": {"$cond": [{"$eq": ["$action", "approve"]}, 1, 0]}
                },
                "first_action": {"$min": "$timestamp"},
                "last_action": {"$max": "$timestamp"},
            }
        },
        {
            "$lookup": {
                "from": "users",
                "localField": "_id",
                "foreignField": "_id",
                "as": "user_info",
            }
        },
        {"$unwind": {"path": "$user_info", "preserveNullAndEmptyArrays": True}},
        {
            "$project": {
                "_id": 0,
                "user_id": "$_id",
                "username": {"$ifNull": ["$user_info.username", "Unknown"]},
                "role": {"$ifNull": ["$user_info.role", "Unknown"]},
                "total_actions": 1,
                "scans": 1,
                "verifications": 1,
                "approvals": 1,
                "first_action": 1,
                "last_action": 1,
            }
        },
        {"$sort": {"total_actions": -1}},
        {"$limit": 1000},
    ]

    return await db.audit_logs.aggregate(pipeline).to_list(1000)


async def generate_session_history_report(db, filters: ReportFilter) -> list[dict]:
    """Generate session history report data."""
    query: dict[str, Any] = {}

    if filters.status:
        query["status"] = filters.status
    if filters.user_id:
        query["user_id"] = filters.user_id

    date_filter = build_date_filter(filters.date_from, filters.date_to)
    if date_filter:
        query["started_at"] = date_filter

    pipeline = [
        {"$match": query},
        {
            "$lookup": {
                "from": "users",
                "localField": "user_id",
                "foreignField": "_id",
                "as": "user_info",
            }
        },
        {
            "$lookup": {
                "from": "count_lines",
                "localField": "session_id",
                "foreignField": "session_id",
                "as": "lines",
            }
        },
        {"$unwind": {"path": "$user_info", "preserveNullAndEmptyArrays": True}},
        {
            "$project": {
                "_id": 0,
                "session_id": 1,
                "username": {"$ifNull": ["$user_info.username", "Unknown"]},
                "rack_id": 1,
                "floor": 1,
                "status": 1,
                "started_at": 1,
                "completed_at": 1,
                "duration_minutes": {
                    "$cond": {
                        "if": {"$and": ["$started_at", "$completed_at"]},
                        "then": {
                            "$divide": [
                                {"$subtract": ["$completed_at", "$started_at"]},
                                60000,
                            ]
                        },
                        "else": None,
                    }
                },
                "items_scanned": {"$size": "$lines"},
                "items_verified": {
                    "$size": {
                        "$filter": {
                            "input": "$lines",
                            "as": "line",
                            "cond": {"$eq": ["$$line.verified", True]},
                        }
                    }
                },
            }
        },
        {"$sort": {"started_at": -1}},
        {"$limit": 5000},
    ]

    return await db.verification_sessions.aggregate(pipeline).to_list(5000)


async def generate_audit_trail_report(db, filters: ReportFilter) -> list[dict]:
    """Generate audit trail report data."""
    query: dict[str, Any] = {}

    if filters.user_id:
        query["user_id"] = filters.user_id

    date_filter = build_date_filter(filters.date_from, filters.date_to)
    if date_filter:
        query["timestamp"] = date_filter

    pipeline = [
        {"$match": query},
        {
            "$lookup": {
                "from": "users",
                "localField": "user_id",
                "foreignField": "_id",
                "as": "user_info",
            }
        },
        {"$unwind": {"path": "$user_info", "preserveNullAndEmptyArrays": True}},
        {
            "$project": {
                "_id": 0,
                "timestamp": 1,
                "action": 1,
                "user_id": 1,
                "username": {"$ifNull": ["$user_info.username", "System"]},
                "role": {"$ifNull": ["$user_info.role", "system"]},
                "target_type": 1,
                "target_id": 1,
                "details": 1,
                "ip_address": 1,
            }
        },
        {"$sort": {"timestamp": -1}},
        {"$limit": 10000},
    ]

    return await db.audit_logs.aggregate(pipeline).to_list(10000)


# Report Generator Dispatch
REPORT_GENERATORS = {
    "stock_summary": generate_stock_summary,
    "variance_report": generate_variance_report,
    "user_activity": generate_user_activity_report,
    "session_history": generate_session_history_report,
    "audit_trail": generate_audit_trail_report,
}


# API Endpoints
@report_generation_router.get("/types")
async def get_report_types(current_user: dict = Depends(get_current_user)):
    """Get available report types with descriptions."""
    return {
        "report_types": [
            {"id": key, "name": value["name"], "description": value["description"]}
            for key, value in REPORT_TYPES.items()
        ]
    }


@report_generation_router.post("/generate", response_model=ReportResponse)
async def generate_report(
    request: ReportRequest,
    current_user: dict = Depends(require_role("admin", "supervisor")),
):
    """
    Generate a report with specified filters.
    Returns data in JSON format by default.
    """
    if request.report_type not in REPORT_TYPES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid report type. Valid types: {list(REPORT_TYPES.keys())}",
        )

    db = get_db()
    filters = request.filters or ReportFilter()

    # Generate report data
    generator = REPORT_GENERATORS[request.report_type]
    try:
        data = await generator(db, filters)
    except Exception as e:
        logger.error(f"Error generating report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate report",
        )

    # Build summary
    filters_applied = {k: v for k, v in filters.dict().items() if v is not None}

    summary = ReportSummary(
        total_records=len(data),
        generated_at=datetime.utcnow().isoformat(),
        filters_applied=filters_applied,
        report_type=request.report_type,
        report_name=REPORT_TYPES[request.report_type]["name"],
    )

    return ReportResponse(summary=summary, data=data)


@report_generation_router.post("/export/csv")
async def export_report_csv(
    request: ReportRequest,
    current_user: dict = Depends(require_role("admin", "supervisor")),
):
    """
    Export report as CSV file.
    """
    if request.report_type not in REPORT_TYPES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid report type"
        )

    db = get_db()
    filters = request.filters or ReportFilter()

    # Generate report data
    generator = REPORT_GENERATORS[request.report_type]
    try:
        data = await generator(db, filters)
    except Exception as e:
        logger.error(f"Error generating report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate report",
        )

    if not data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No data found for the specified filters",
        )

    # Create CSV
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()

    for row in data:
        sanitized_row = {k: sanitize_for_csv(v) for k, v in row.items()}
        writer.writerow(sanitized_row)

    output.seek(0)

    filename = f"{request.report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@report_generation_router.post("/export/xlsx")
async def export_report_xlsx(
    request: ReportRequest,
    current_user: dict = Depends(require_role("admin", "supervisor")),
):
    """
    Export report as Excel XLSX file.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Excel export not available. Install openpyxl package.",
        )

    if request.report_type not in REPORT_TYPES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid report type"
        )

    db = get_db()
    filters = request.filters or ReportFilter()

    generator = REPORT_GENERATORS[request.report_type]
    try:
        data = await generator(db, filters)
    except Exception as e:
        logger.error(f"Error generating report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate report",
        )

    if not data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No data found for the specified filters",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = REPORT_TYPES[request.report_type]["name"][:31]

    headers = list(data[0].keys())
    _write_xlsx_headers(ws, headers)
    _write_xlsx_data(ws, data, headers)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{request.report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@report_generation_router.get("/filters/{report_type}")
async def get_report_filter_options(
    report_type: str, current_user: dict = Depends(get_current_user)
):
    """
    Get available filter options for a specific report type.
    Returns distinct values for filterable fields.
    """
    if report_type not in REPORT_TYPES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid report type"
        )

    db = get_db()

    try:
        # Get distinct values for common filters
        warehouses = await db.erp_items.distinct("warehouse")
        floors = await db.erp_items.distinct("floor")
        categories = await db.erp_items.distinct("category")
        statuses = await db.verification_records.distinct("status")

        # Get user list for admin/supervisor
        users = []
        if current_user.get("role") in ["admin", "supervisor"]:
            user_cursor = db.users.find({}, {"_id": 1, "username": 1, "role": 1})
            users = [
                {
                    "id": str(u["_id"]),
                    "username": u["username"],
                    "role": u.get("role", "staff"),
                }
                async for u in user_cursor
            ]

        return {
            "report_type": report_type,
            "filters": {
                "warehouses": warehouses,
                "floors": floors,
                "categories": categories,
                "statuses": statuses,
                "users": users,
            },
        }

    except Exception as e:
        logger.error(f"Error fetching filter options: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch filter options",
        )
